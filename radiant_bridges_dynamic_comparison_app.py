import io
import json
from datetime import date
from typing import Any, Dict, List

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

try:
    import pdfplumber
    PDF_READ_AVAILABLE = True
except Exception:
    PDF_READ_AVAILABLE = False


st.set_page_config(
    page_title="Radiant Bridges Dynamic Comparison App",
    page_icon="🛗",
    layout="wide",
)

SPEC_LIST = [
    "CAPACITY", "SPEED", "DOOR TYPE", "DOOR SIZE (W X H)",
    "SHAFT SIZE (W X D)", "CABIN SIZE (W X D X H)",
    "OVER HEAD HEIGHT", "PIT DEPTH", "NO. OF LIFTS",
    "Lift code", "Machine location", "Operation", "No. of Stops",
    "Travel Height", "Car wall", "Front Wall", "Ceiling", "Mirror",
    "Hand rail", "Skirting", "Decoration", "Door Material",
    "Sill Material", "COP Panel", "LOP",
    "Landing Jamb In Ground Floor", "Landing Jamb In Other Floors",
    "Hall Indicator", "Made", "Remarks"
]

DEFAULT_VENDORS = ["KONE", "TKE", "EEE", "AG MELCO"]

DEFAULT_PROJECT_INFO = {
    "project": "RADIANT BRIDGES PROJECT",
    "document_title": "ELEVATOR TECHNICAL COMPARISON",
    "client": "Radiant Real Estate",
    "main_contractor": "ATGC Transport & General Contracting L.L.C.-SPC",
    "material_work": "Elevators",
    "pr_no": "",
    "revision": "Rev. 0",
    "comparison_date": date.today().isoformat(),
}


def make_default_df(vendors: List[str]) -> pd.DataFrame:
    columns = ["Specification", "Consultant"] + vendors
    df = pd.DataFrame("", index=range(len(SPEC_LIST)), columns=columns)
    df["Specification"] = SPEC_LIST
    return df


def sync_vendor_columns(df: pd.DataFrame, vendors: List[str]) -> pd.DataFrame:
    required_cols = ["Specification", "Consultant"] + vendors

    if "Specification" not in df.columns:
        df.insert(0, "Specification", "")

    if "Consultant" not in df.columns:
        df.insert(1, "Consultant", "")

    for col in required_cols:
        if col not in df.columns:
            df[col] = ""

    return df[required_cols].fillna("").copy()


def init_state() -> None:
    if "vendors" not in st.session_state:
        st.session_state.vendors = DEFAULT_VENDORS.copy()

    if "project_info" not in st.session_state:
        st.session_state.project_info = DEFAULT_PROJECT_INFO.copy()

    if "groups" not in st.session_state:
        st.session_state.groups = {
            "Tower A": {
                "PL1 & PL2": make_default_df(st.session_state.vendors)
            }
        }

    if "active_tower" not in st.session_state:
        st.session_state.active_tower = "Tower A"

    if "active_group" not in st.session_state:
        st.session_state.active_group = "PL1 & PL2"

    if "attachments" not in st.session_state:
        st.session_state.attachments = []


def sync_all_groups() -> None:
    for tower in st.session_state.groups:
        for group in st.session_state.groups[tower]:
            st.session_state.groups[tower][group] = sync_vendor_columns(
                st.session_state.groups[tower][group],
                st.session_state.vendors,
            )


def ensure_active_selection() -> None:
    if not st.session_state.groups:
        st.session_state.active_tower = ""
        st.session_state.active_group = ""
        return

    if st.session_state.active_tower not in st.session_state.groups:
        st.session_state.active_tower = list(st.session_state.groups.keys())[0]

    if not st.session_state.groups[st.session_state.active_tower]:
        st.session_state.active_group = ""
        return

    if st.session_state.active_group not in st.session_state.groups[st.session_state.active_tower]:
        st.session_state.active_group = list(
            st.session_state.groups[st.session_state.active_tower].keys()
        )[0]


def clean_text(value: Any) -> str:
    return str(value).replace("\n", " ").strip()


def read_uploaded_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()

    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file).fillna("")

    if name.endswith((".xlsx", ".xls")):
        return pd.read_excel(uploaded_file).fillna("")

    if name.endswith(".pdf"):
        if not PDF_READ_AVAILABLE:
            return pd.DataFrame({"PDF Error": ["pdfplumber not installed. Add pdfplumber to requirements.txt"]})

        rows = []
        with pdfplumber.open(uploaded_file) as pdf:
            for page_no, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                for table in tables or []:
                    for row in table:
                        rows.append(row)

                text = page.extract_text()
                if text:
                    for line in text.split("\n"):
                        rows.append([f"Page {page_no}", line])

        if not rows:
            return pd.DataFrame({"PDF Content": ["No extractable text/table found"]})

        max_cols = max(len(r) for r in rows)
        rows = [r + [""] * (max_cols - len(r)) for r in rows]
        return pd.DataFrame(rows, columns=[f"Column {i+1}" for i in range(max_cols)]).fillna("")

    raise ValueError("Unsupported file format")


def build_excel(groups, vendors, project_info) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    group_fill = PatternFill("solid", fgColor="BDD7EE")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    spec_fill = PatternFill("solid", fgColor="E2F0D9")
    label_fill = PatternFill("solid", fgColor="F2F2F2")

    columns = ["Specification", "Consultant"] + vendors
    last_col = len(columns)
    row = 1

    def style_range(r1, c1, r2, c2, fill=None, bold=False, color="000000", align="center"):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(r, c)
                cell.border = border
                cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
                cell.font = Font(bold=bold, color=color)
                if fill:
                    cell.fill = fill

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ws.cell(row, 1, project_info.get("document_title", "ELEVATOR TECHNICAL COMPARISON"))
    style_range(row, 1, row, last_col, title_fill, True)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ws.cell(row, 1, project_info.get("project", "RADIANT BRIDGES PROJECT"))
    style_range(row, 1, row, last_col, title_fill, True)
    row += 2

    info_rows = [
        ("Client", project_info.get("client", ""), "Date", project_info.get("comparison_date", "")),
        ("Main Contractor", project_info.get("main_contractor", ""), "Revision", project_info.get("revision", "")),
        ("Material / Work", project_info.get("material_work", ""), "PR No.", project_info.get("pr_no", "")),
    ]

    mid = max(2, last_col // 2)

    for a, b, c, d in info_rows:
        ws.cell(row, 1, a)
        ws.cell(row, 2, b)
        ws.cell(row, mid + 1, c)
        ws.cell(row, mid + 2, d)
        style_range(row, 1, row, last_col, None, False, align="left")
        ws.cell(row, 1).fill = label_fill
        ws.cell(row, mid + 1).fill = label_fill
        row += 1

    row += 1

    for tower, group_data in groups.items():
        for group, df in group_data.items():
            df = sync_vendor_columns(df, vendors)

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
            ws.cell(row, 1, f"{tower} - {group}")
            style_range(row, 1, row, last_col, group_fill, True)
            row += 1

            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row, col_idx, col_name)
            style_range(row, 1, row, last_col, header_fill, True, color="FFFFFF")
            row += 1

            for _, data_row in df.iterrows():
                for col_idx, col_name in enumerate(columns, start=1):
                    ws.cell(row, col_idx, str(data_row.get(col_name, "")))
                    ws.cell(row, col_idx).border = border
                    ws.cell(row, col_idx).alignment = Alignment(wrap_text=True, vertical="top")
                    if col_idx == 1:
                        ws.cell(row, col_idx).fill = spec_fill
                        ws.cell(row, col_idx).font = Font(bold=True)
                row += 1

            row += 2

    for col in range(1, last_col + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "Z"].width = 24
    ws.column_dimensions["A"].width = 32

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_simple_pdf(groups, vendors, project_info) -> bytes:
    lines = []
    lines.append(project_info.get("document_title", "ELEVATOR TECHNICAL COMPARISON"))
    lines.append(project_info.get("project", "RADIANT BRIDGES PROJECT"))
    lines.append("")
    lines.append(f"Client: {project_info.get('client', '')}")
    lines.append(f"Main Contractor: {project_info.get('main_contractor', '')}")
    lines.append(f"Material / Work: {project_info.get('material_work', '')}")
    lines.append(f"Date: {project_info.get('comparison_date', '')}")
    lines.append("")

    for tower, group_data in groups.items():
        for group, df in group_data.items():
            df = sync_vendor_columns(df, vendors)

            lines.append("=" * 90)
            lines.append(f"CONSULTANT SPECIFICATION - {tower} - {group}")
            lines.append("=" * 90)

            for _, r in df.iterrows():
                lines.append(f"{r.get('Specification', '')}: {r.get('Consultant', '')}")

            lines.append("")

            for vendor in vendors:
                lines.append("-" * 90)
                lines.append(f"VENDOR OFFER - {vendor} - {tower} - {group}")
                lines.append("-" * 90)

                for _, r in df.iterrows():
                    lines.append(f"{r.get('Specification', '')}: {r.get(vendor, '')}")

                lines.append("")

    return make_pdf_from_lines(lines)


def make_pdf_from_lines(lines: List[str]) -> bytes:
    def esc(s: str) -> str:
        s = str(s).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        return s[:120]

    pages = []
    chunk_size = 42
    for i in range(0, len(lines), chunk_size):
        pages.append(lines[i:i + chunk_size])

    objects = []
    objects.append("<< /Type /Catalog /Pages 2 0 R >>")
    objects.append(f"<< /Type /Pages /Kids [{' '.join(f'{3 + i*2} 0 R' for i in range(len(pages)))}] /Count {len(pages)} >>")

    for page_index, page_lines in enumerate(pages):
        page_obj = 3 + page_index * 2
        content_obj = page_obj + 1

        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 842 595] "
            f"/Resources << /Font << /F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> >> >> "
            f"/Contents {content_obj} 0 R >>"
        )

        y = 560
        content = ["BT /F1 9 Tf"]
        for line in page_lines:
            content.append(f"40 {y} Td ({esc(line)}) Tj")
            content.append("0 -12 Td")
            y -= 12
        content.append("ET")
        stream = "\n".join(content)
        objects.append(f"<< /Length {len(stream.encode('latin-1', 'ignore'))} >>\nstream\n{stream}\nendstream")

    pdf = "%PDF-1.4\n"
    offsets = [0]

    for i, obj in enumerate(objects, start=1):
        offsets.append(len(pdf.encode("latin-1")))
        pdf += f"{i} 0 obj\n{obj}\nendobj\n"

    xref_pos = len(pdf.encode("latin-1"))
    pdf += f"xref\n0 {len(objects) + 1}\n"
    pdf += "0000000000 65535 f \n"

    for off in offsets[1:]:
        pdf += f"{off:010d} 00000 n \n"

    pdf += f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF"

    return pdf.encode("latin-1", "ignore")


def export_backup() -> bytes:
    data = {
        "vendors": st.session_state.vendors,
        "project_info": st.session_state.project_info,
        "groups": {
            tower: {
                group: df.to_dict(orient="records")
                for group, df in group_data.items()
            }
            for tower, group_data in st.session_state.groups.items()
        },
    }
    return json.dumps(data, indent=2).encode("utf-8")


init_state()
sync_all_groups()
ensure_active_selection()

with st.sidebar:
    st.header("⚙️ Project Info")
    pi = st.session_state.project_info

    pi["project"] = st.text_input("Project", pi.get("project", ""))
    pi["document_title"] = st.text_input("Document Title", pi.get("document_title", ""))
    pi["client"] = st.text_input("Client", pi.get("client", ""))
    pi["main_contractor"] = st.text_input("Main Contractor", pi.get("main_contractor", ""))
    pi["material_work"] = st.text_input("Material / Work", pi.get("material_work", ""))
    pi["pr_no"] = st.text_input("PR No.", pi.get("pr_no", ""))
    pi["revision"] = st.text_input("Revision", pi.get("revision", ""))
    pi["comparison_date"] = st.text_input("Date", pi.get("comparison_date", ""))

    st.divider()
    st.header("🏢 Structure")

    if st.session_state.groups:
        tower_list = list(st.session_state.groups.keys())
        st.session_state.active_tower = st.selectbox("Tower / Section", tower_list)

        group_list = list(st.session_state.groups[st.session_state.active_tower].keys())
        if group_list:
            st.session_state.active_group = st.selectbox("Group", group_list)
        else:
            st.session_state.active_group = ""

    new_tower = st.text_input("Add Tower / Section")
    if st.button("Add Tower / Section"):
        name = new_tower.strip()
        if name and name not in st.session_state.groups:
            st.session_state.groups[name] = {}
            st.session_state.active_tower = name
            st.rerun()

    new_group = st.text_input("Add Group")
    if st.button("Add Group"):
        name = new_group.strip()
        if st.session_state.active_tower and name:
            st.session_state.groups[st.session_state.active_tower][name] = make_default_df(st.session_state.vendors)
            st.session_state.active_group = name
            st.rerun()

    c1, c2 = st.columns(2)
    if c1.button("Remove Group"):
        if st.session_state.active_tower and st.session_state.active_group:
            del st.session_state.groups[st.session_state.active_tower][st.session_state.active_group]
            ensure_active_selection()
            st.rerun()

    if c2.button("Remove Tower"):
        if st.session_state.active_tower:
            del st.session_state.groups[st.session_state.active_tower]
            ensure_active_selection()
            st.rerun()

    st.divider()
    st.header("🏷️ Vendors")

    new_vendor = st.text_input("Add Vendor")
    if st.button("Add Vendor"):
        name = new_vendor.strip().upper()
        if name and name not in st.session_state.vendors:
            st.session_state.vendors.append(name)
            sync_all_groups()
            st.rerun()

    for vendor in list(st.session_state.vendors):
        a, b = st.columns([4, 1])
        a.write(vendor)
        if b.button("❌", key=f"del_{vendor}"):
            st.session_state.vendors.remove(vendor)
            sync_all_groups()
            st.rerun()


st.title("🛗 Radiant Bridges Dynamic Comparison App")
st.caption("No fpdf or xlsxwriter required. Excel uses openpyxl. PDF uses built-in simple PDF generator.")

m1, m2, m3, m4 = st.columns(4)
m1.metric("Towers", len(st.session_state.groups))
m2.metric("Groups", sum(len(x) for x in st.session_state.groups.values()))
m3.metric("Vendors", len(st.session_state.vendors))
m4.metric("Active", f"{st.session_state.active_tower or '-'} / {st.session_state.active_group or '-'}")

st.divider()

if not st.session_state.groups:
    st.warning("No tower/section available. Add one from sidebar.")
    st.stop()

if not st.session_state.active_tower or not st.session_state.groups.get(st.session_state.active_tower):
    st.warning("Please add at least one group under selected tower/section.")
    st.stop()

active_tower = st.session_state.active_tower
active_group = st.session_state.active_group

st.subheader(f"📊 {active_tower} > {active_group}")

current_df = sync_vendor_columns(
    st.session_state.groups[active_tower][active_group],
    st.session_state.vendors,
)

edited_df = st.data_editor(
    current_df,
    num_rows="dynamic",
    use_container_width=True,
    hide_index=True,
    key=f"editor_{active_tower}_{active_group}",
)

st.session_state.groups[active_tower][active_group] = sync_vendor_columns(
    edited_df,
    st.session_state.vendors,
)

st.divider()
st.subheader("📎 Attachments + Mapping")

attachment_uploads = st.file_uploader(
    "Upload attachments / vendor offers / consultant specs",
    type=["pdf", "xlsx", "xls", "csv", "docx", "png", "jpg", "jpeg"],
    accept_multiple_files=True,
)

if attachment_uploads:
    existing = {x["name"] for x in st.session_state.attachments}
    for file in attachment_uploads:
        if file.name not in existing:
            st.session_state.attachments.append({
                "name": file.name,
                "type": file.type,
                "size": file.size,
                "bytes": file.getvalue(),
            })

if st.session_state.attachments:
    for i, item in enumerate(list(st.session_state.attachments)):
        a, b, c = st.columns([5, 2, 1])
        a.write(f"📄 {item['name']}")
        b.write(f"{round(item['size'] / 1024, 1)} KB")
        if c.button("Remove", key=f"remove_{i}"):
            st.session_state.attachments.pop(i)
            st.rerun()

    selected_name = st.selectbox(
        "Select attachment for mapping",
        [x["name"] for x in st.session_state.attachments],
    )

    selected = next(x for x in st.session_state.attachments if x["name"] == selected_name)
    file_obj = io.BytesIO(selected["bytes"])
    file_obj.name = selected["name"]

    if selected_name.lower().endswith((".xlsx", ".xls", ".csv", ".pdf")):
        try:
            raw_df = read_uploaded_file(file_obj)
            st.dataframe(raw_df.head(20), use_container_width=True, hide_index=True)

            target_cols = ["Consultant"] + st.session_state.vendors
            options = ["-- Do not import --"] + list(raw_df.columns)

            mapping = {}
            cols = st.columns(min(3, len(target_cols)))

            for i, target in enumerate(target_cols):
                with cols[i % len(cols)]:
                    mapping[target] = st.selectbox(
                        f"Map to {target}",
                        options,
                        key=f"map_{selected_name}_{target}",
                    )

            if st.button("Apply Mapping", type="primary"):
                df = st.session_state.groups[active_tower][active_group].copy()
                max_rows = min(len(df), len(raw_df))

                for target, source in mapping.items():
                    if source != "-- Do not import --":
                        df.loc[:max_rows - 1, target] = raw_df[source].astype(str).head(max_rows).tolist()

                st.session_state.groups[active_tower][active_group] = sync_vendor_columns(
                    df,
                    st.session_state.vendors,
                )
                st.success("Mapping applied.")
                st.rerun()

        except Exception as exc:
            st.error(f"Unable to map file: {exc}")
    else:
        st.info("This file is stored as attachment only. Mapping supports Excel, CSV, and PDF.")
else:
    st.info("No attachments uploaded.")

st.divider()
st.subheader("📤 Export")

excel_bytes = build_excel(
    st.session_state.groups,
    st.session_state.vendors,
    st.session_state.project_info,
)

pdf_bytes = build_simple_pdf(
    st.session_state.groups,
    st.session_state.vendors,
    st.session_state.project_info,
)

x1, x2, x3 = st.columns(3)

with x1:
    st.download_button(
        "📥 Download Excel",
        data=excel_bytes,
        file_name="RADIANT_BRIDGES_COMPARISON.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

with x2:
    st.download_button(
        "📄 Download PDF",
        data=pdf_bytes,
        file_name="RADIANT_BRIDGES_COMPARISON.pdf",
        mime="application/pdf",
        use_container_width=True,
    )

with x3:
    st.download_button(
        "💾 Backup JSON",
        data=export_backup(),
        file_name="radiant_bridges_backup.json",
        mime="application/json",
        use_container_width=True,
    )

with st.expander("Install / Run"):
    st.code(
        """pip install streamlit pandas openpyxl pdfplumber
streamlit run radiant_bridges_dynamic_comparison_app.py""",
        language="bash",
    )
